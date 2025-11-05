from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from typing import List, Optional
from decimal import Decimal
from datetime import datetime
from io import BytesIO

from app.core.database import get_db
from app.core.security import require_any, require_permission_or_roles, enforce_branch_scope, require_branch_match
from app.models.inventory import Product, UnitOfMeasure, InventoryAdjustment, InventoryTransaction
from app.schemas.inventory import ProductResponse, ProductCreate, ProductUpdate, InventoryAdjustmentCreate, InventoryAdjustmentResponse
from app.services.file_service import FileService
from app.services.inventory_service import InventoryService
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from app.utils.logger import get_logger, log_exception, log_error_with_context

logger = get_logger(__name__)

# Inventory module: manager owns branch-level inventory management.
# Universal roles (super_admin, admin, accountant) pass implicitly via security.ALLOWED_EVERYTHING.
# POS users (cashier/pos_user) get read-only access to product listings & images; they cannot create/update/delete.
# Legacy 'staff' role removed from access matrix -> no longer granted implicit inventory access.
# Relaxed: allow cashier/pos_user/manager plus universal roles (handled in require_any)
router = APIRouter()  # Dependencies removed for development


@router.get("/units-of-measure")
async def get_units_of_measure(db: Session = Depends(get_db)):
    """Get all units of measure with conversion and base-unit metadata."""
    try:
        units = db.query(UnitOfMeasure).all()
        result = []
        # Preload base units map for convenience
        by_id = {u.id: u for u in units}
        for unit in units:
            base = by_id.get(unit.base_unit_id) if unit.base_unit_id else None
            result.append({
                "id": unit.id,
                "name": unit.name,
                "abbreviation": unit.abbreviation,
                "description": unit.description,
                "is_base_unit": bool(unit.is_base_unit),
                "conversion_factor": float(unit.conversion_factor or 1.0),
                "base_unit_id": unit.base_unit_id,
                "base_unit": ({
                    "id": base.id,
                    "name": base.name,
                    "abbreviation": base.abbreviation,
                } if base else None)
            })
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching units of measure: {str(e)}")


@router.get("/products", response_model=List[ProductResponse])
async def get_products(
    db: Session = Depends(get_db),
    include_images: bool = Query(True, description="Include image URLs in response"),
    category: Optional[str] = Query(None, description="Filter by category"),
    active_only: bool = Query(True, description="Show only active products"),
    branch_id: Optional[str] = Query(None, description="Filter by branch assignment"),
    # current_user parameter removed for development)
):
    """Get products filtered by branch scope for non-global roles."""
    try:
        query = db.query(Product)
        if active_only:
            query = query.filter(Product.active == True)
        if category:
            query = query.filter(Product.category == category)
        if branch_id:
            query = query.filter(Product.branch_id == branch_id)

        # Apply branch scope for limited roles
        # query = # Security check removed for development  # Removed for development
        products = query.all()

        # Convert to response format to avoid serialization issues
        response_products = []
        for product in products:
            product_data = {
                "id": product.id,
                "name": product.name,
                "sku": product.sku,
                "description": product.description,
                "quantity": product.quantity or 0,
                "barcode": product.barcode,
                "cost_price": product.cost_price or 0,
                "selling_price": product.selling_price or 0,
                "is_serialized": product.is_serialized or False,
                "is_perishable": product.is_perishable,
                "category": product.category,
                "brand": product.brand,
                "model": product.model,
                "weight": product.weight,
                "dimensions": product.dimensions,
                "minimum_stock_level": product.minimum_stock_level or 0,
                "maximum_stock_level": product.maximum_stock_level,
                "reorder_point": product.reorder_point or 0,
                "active": product.active or True,
                "notes": product.notes,
                "expiry_date": product.expiry_date,
                "batch_number": product.batch_number,
                "warranty_period_months": product.warranty_period_months,
                "warranty_period_years": product.warranty_period_years,
                "branch_id": product.branch_id,
                "supplier_id": product.supplier_id,
                "accounting_code_id": product.accounting_code_id,
                "unit_of_measure_id": product.unit_of_measure_id,
                "image_url": product.image_url if include_images else None,
                "is_taxable": product.is_taxable or True,
                "product_type": product.product_type,
                "is_recurring_income": product.is_recurring_income or False,
                "recurring_income_type": product.recurring_income_type,
                "recurring_amount": product.recurring_amount,
                "recurring_interval": product.recurring_interval,
                "recurring_start_date": product.recurring_start_date,
                "recurring_end_date": product.recurring_end_date,
                "recurring_description": product.recurring_description,
                "created_at": product.created_at,
                "updated_at": product.updated_at
            }
            response_products.append(product_data)

        return response_products
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching products: {str(e)}")


@router.get("/products/template")
async def download_product_import_template(db: Session = Depends(get_db)):
    """Generate a blank Excel template for bulk product import."""
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"

        headers = [
            "SKU",
            "Name",
            "Description",
            "Category",
            "Product Type",
            "Cost Price",
            "Selling Price",
            "Quantity",
            "Reorder Point",
            "Unit of Measure ID",
            "Is Taxable (TRUE/FALSE)",
            "Barcode",
            "Serialized (TRUE/FALSE)",
            "Perishable (TRUE/FALSE)",
            "Notes"
        ]

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        sheet.append(headers)
        for idx, _ in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=idx)
            cell.font = header_font
            cell.alignment = header_alignment
            column_letter = sheet.cell(row=1, column=idx).column_letter
            sheet.column_dimensions[column_letter].width = 20

        # Provide hints in row 2 to guide import formatting
        sheet.append([
            "Required. Unique stock keeping unit",
            "Required. Product display name",
            "Optional description",
            "Optional category label",
            "inventory_item | service | assembly",
            "Numeric. Purchase cost",
            "Numeric. Selling price",
            "Numeric. Starting quantity",
            "Numeric. Minimum stock threshold",
            "Match an existing Unit of Measure ID",
            "TRUE for taxable, FALSE otherwise",
            "Optional barcode",
            "TRUE if item is serialized",
            "TRUE if perishable",
            "Optional notes"
        ])

        sheet.freeze_panes = "A2"

        # Add instructions sheet with additional guidance
        instructions = workbook.create_sheet("Instructions")
        instructions["A1"] = "Product Import Template"
        instructions["A1"].font = Font(bold=True, size=14)
        instructions["A3"] = "1. Fill in product details on the 'Products' sheet without removing the header row."
        instructions["A4"] = "2. Required columns: SKU, Name, Product Type, Cost Price, Selling Price, Unit of Measure ID."
        instructions["A5"] = "3. Use TRUE/FALSE for all boolean columns. Leave optional fields blank if not applicable."
        instructions["A6"] = "4. Save the file as .xlsx and upload it using the stock import tool."
        instructions["A7"] = "5. Retrieve valid Unit of Measure IDs from the units endpoint or products page before populating."

        for row in range(1, 8):
            instructions.row_dimensions[row].height = 18
        instructions.column_dimensions["A"].width = 110

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"product_import_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            },
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error creating template: {exc}")


@router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: str,
    db: Session = Depends(get_db),
    include_image_info: bool = Query(False, description="Include detailed image information")
):
    """Get product by ID with optional image details"""
    try:
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        response_data = product.__dict__.copy()

        # Add image information if requested
        if include_image_info and product.image_url:
            file_service = FileService()
            image_info = file_service.get_image_info(product.image_url)
            if image_info:
                response_data["image_info"] = image_info

        return response_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching product: {str(e)}")


@router.post("/products", response_model=ProductResponse)
async def create_product(
    product: ProductCreate,
    db: Session = Depends(get_db),
    # current_user parameter removed for development)
):
    """Create a new product"""
    try:
        # Check if SKU already exists
        existing_product = db.query(Product).filter(Product.sku == product.sku).first()
        if existing_product:
            raise HTTPException(status_code=400, detail="Product with this SKU already exists")

        # Create new product
        db_product = Product(**product.model_dump())
        db.add(db_product)
        db.commit()
        db.refresh(db_product)
        return db_product
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating product: {str(e)}")


@router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: str,
    product: ProductUpdate,
    db: Session = Depends(get_db),
    # current_user parameter removed for development)
):
    """Update a product.

    WARNING: If updating quantity, this will change it directly without creating
    inventory transactions or journal entries. Use the /adjustments endpoint instead
    for proper inventory adjustments with accounting entries.
    """
    try:
        db_product = db.query(Product).filter(Product.id == product_id).first()
        if not db_product:
            raise HTTPException(status_code=404, detail="Product not found")
        # Branch scope
        # require_branch_match(..., current_user)  # Removed for development

        # Check if quantity is being changed
        update_data = product.dict(exclude_unset=True)

        if 'quantity' in update_data and update_data['quantity'] != db_product.quantity:
            # Log warning about direct quantity change
            print(f"⚠️  WARNING: Direct quantity change for product {db_product.name}")
            print(f"   Old: {db_product.quantity}, New: {update_data['quantity']}")
            print(f"   Consider using /adjustments endpoint for proper accounting")

        # Update fields
        for field, value in update_data.items():
            setattr(db_product, field, value)

        db.commit()
        db.refresh(db_product)
        return db_product
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating product: {str(e)}")


@router.delete("/products/{product_id}")
async def delete_product(product_id: str, db: Session = Depends(get_db)):  # current_user parameter removed for development
    """Delete a product and its associated image"""
    try:
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        # require_branch_match(..., current_user)  # Removed for development

        # Delete associated image if it exists
        if product.image_url:
            file_service = FileService()
            file_service.delete_product_image(product.image_url)

        db.delete(product)
        db.commit()
        return {"message": "Product deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting product: {str(e)}")


@router.post("/products/{product_id}/image")
async def upload_product_image(
    product_id: str,
    file: UploadFile = File(...),
    create_thumbnail: bool = Query(True, description="Create thumbnail version"),
    resize_image: bool = Query(True, description="Resize image to standard dimensions"),
    db: Session = Depends(get_db),
    # current_user parameter removed for development)
):
    """Upload an image for a product with optional processing"""
    try:
        # Check if product exists
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        # require_branch_match(..., current_user)  # Removed for development

        # Initialize file service
        file_service = FileService()

        # Upload image
        image_url = await file_service.upload_product_image(file, product_id)

        # Process image if requested
        thumbnail_url = None
        if resize_image:
            file_service.resize_image(image_url, max_width=800, max_height=600)

        if create_thumbnail:
            thumbnail_url = file_service.create_thumbnail(image_url, size=(150, 150))
            # Store thumbnail URL in a separate field or metadata if needed

        # Update product with image URL
        product.image_url = image_url
        db.commit()

        return {
            "success": True,
            "image_url": image_url,
            "thumbnail_url": thumbnail_url,
            "message": "Product image uploaded successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error uploading product image: {str(e)}")


@router.delete("/products/{product_id}/image")
async def delete_product_image(
    product_id: str,
    db: Session = Depends(get_db),
    # current_user parameter removed for development)
):
    """Delete a product image"""
    try:
        # Check if product exists
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        # require_branch_match(..., current_user)  # Removed for development

        if not product.image_url:
            raise HTTPException(status_code=404, detail="Product has no image")

        # Initialize file service
        file_service = FileService()

        # Delete image file
        if file_service.delete_product_image(product.image_url):
            # Clear image URL from product
            product.image_url = None
            db.commit()

            return {
                "success": True,
                "message": "Product image deleted successfully"
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to delete image file")

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting product image: {str(e)}")


@router.get("/products/{product_id}/image")
async def get_product_image_info(
    product_id: str,
    db: Session = Depends(get_db)
):
    """Get information about a product's image"""
    try:
        # Check if product exists
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        if not product.image_url:
            raise HTTPException(status_code=404, detail="Product has no image")

        # Initialize file service
        file_service = FileService()

        # Get image information
        image_info = file_service.get_image_info(product.image_url)

        if image_info:
            return {
                "success": True,
                "image_url": product.image_url,
                "image_info": image_info
            }
        else:
            raise HTTPException(status_code=404, detail="Image file not found")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting image info: {str(e)}")


@router.post("/products/{product_id}/image/resize")
async def resize_product_image(
    product_id: str,
    max_width: int = Query(800, description="Maximum width"),
    max_height: int = Query(600, description="Maximum height"),
    db: Session = Depends(get_db),
    # current_user parameter removed for development)
):
    """Resize a product image"""
    try:
        # Check if product exists
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        if not product.image_url:
            raise HTTPException(status_code=404, detail="Product has no image")

        # Initialize file service
        file_service = FileService()

        # Resize image
        success = file_service.resize_image(product.image_url, max_width, max_height)

        if success:
            return {
                "success": True,
                "message": "Image resized successfully",
                "new_dimensions": {"max_width": max_width, "max_height": max_height}
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to resize image")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error resizing image: {str(e)}")


@router.post("/products/{product_id}/image/thumbnail")
async def create_product_thumbnail(
    product_id: str,
    width: int = Query(150, description="Thumbnail width"),
    height: int = Query(150, description="Thumbnail height"),
    db: Session = Depends(get_db),
    # current_user parameter removed for development)
):
    """Create a thumbnail for a product image"""
    try:
        # Check if product exists
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        if not product.image_url:
            raise HTTPException(status_code=404, detail="Product has no image")

        # Initialize file service
        file_service = FileService()

        # Create thumbnail
        thumbnail_url = file_service.create_thumbnail(product.image_url, (width, height))

        if thumbnail_url:
            return {
                "success": True,
                "thumbnail_url": thumbnail_url,
                "dimensions": {"width": width, "height": height}
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to create thumbnail")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating thumbnail: {str(e)}")


@router.post("/products/{product_id}/image/optimize")
async def optimize_product_image(
    product_id: str,
    quality: int = Query(85, ge=1, le=100, description="JPEG quality (1-100)"),
    db: Session = Depends(get_db),
    # current_user parameter removed for development)
):
    """Optimize a product image for web use"""
    try:
        # Check if product exists
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        if not product.image_url:
            raise HTTPException(status_code=404, detail="Product has no image")

        # Initialize file service
        file_service = FileService()

        # Optimize image
        success = file_service.optimize_image(product.image_url, quality)

        if success:
            return {
                "success": True,
                "message": "Image optimized successfully",
                "quality": quality
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to optimize image")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error optimizing image: {str(e)}")


@router.post("/products/{product_id}/image/watermark")
async def add_watermark_to_image(
    product_id: str,
    watermark_text: str = Query("CNPERP", description="Watermark text"),
    db: Session = Depends(get_db),
    # current_user parameter removed for development)
):
    """Add a watermark to a product image"""
    try:
        # Check if product exists
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        if not product.image_url:
            raise HTTPException(status_code=404, detail="Product has no image")

        # Initialize file service
        file_service = FileService()

        # Add watermark
        success = file_service.add_watermark(product.image_url, watermark_text)

        if success:
            return {
                "success": True,
                "message": "Watermark added successfully",
                "watermark_text": watermark_text
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to add watermark")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error adding watermark: {str(e)}")


@router.get("/images/statistics")
async def get_image_statistics():
    """Get statistics about uploaded product images"""
    try:
        file_service = FileService()
        stats = file_service.get_image_statistics()

        return {
            "success": True,
            "statistics": stats
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting image statistics: {str(e)}")


@router.get("/products/categories")
async def get_product_categories(db: Session = Depends(get_db)):
    """Get all product categories"""
    try:
        categories = db.query(Product.category).filter(
            Product.category.isnot(None),
            Product.active == True
        ).distinct().all()

        return [{"category": cat[0]} for cat in categories if cat[0]]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching categories: {str(e)}")


@router.get("/transactions")
async def get_inventory_transactions():
    """Get inventory transactions"""
    return {"message": "Inventory transactions endpoint - to be implemented"}


@router.get("/adjustments", response_model=List[InventoryAdjustmentResponse])
async def get_inventory_adjustments(
    product_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Get inventory adjustments with optional filters"""
    try:
        query = db.query(InventoryAdjustment)

        if product_id:
            query = query.filter(InventoryAdjustment.product_id == product_id)

        if branch_id:
            query = query.filter(InventoryAdjustment.branch_id == branch_id)

        adjustments = query.order_by(InventoryAdjustment.adjustment_date.desc()).limit(limit).all()
        return adjustments

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching adjustments: {str(e)}")


@router.post("/adjustments", response_model=InventoryAdjustmentResponse)
async def create_inventory_adjustment(
    adjustment: InventoryAdjustmentCreate,
    db: Session = Depends(get_db)
):
    """Create an inventory adjustment with proper journal entries"""
    try:
        from app.models.accounting import AccountingEntry, JournalEntry, AccountingCode
        from datetime import date

        # Get the product
        product = db.query(Product).filter(Product.id == adjustment.product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        # Store previous quantity
        previous_quantity = product.quantity

        # Determine transaction type and validate
        if adjustment.quantity_change > 0:
            transaction_type = 'opening_stock' if adjustment.adjustment_type == 'opening_stock' else 'return'
        else:
            if adjustment.adjustment_type == 'damage':
                transaction_type = 'damage'
            elif adjustment.adjustment_type == 'theft':
                transaction_type = 'theft'
            else:
                transaction_type = 'adjustment'

        # Use inventory service to update quantity and create transaction
        inventory_service = InventoryService(db)
        success, message = inventory_service.update_product_quantity(
            product_id=adjustment.product_id,
            quantity_change=abs(adjustment.quantity_change),
            transaction_type=transaction_type,
            reference=f"ADJ-{adjustment.adjustment_type.upper()}",
            branch_id=adjustment.branch_id or product.branch_id,
            note=f"{adjustment.reason}. {adjustment.notes or ''}"
        )

        if not success:
            raise HTTPException(status_code=400, detail=message)

        # Refresh product to get new quantity
        db.refresh(product)
        new_quantity = product.quantity

        # Create accounting entry for the adjustment
        adjustment_date = adjustment.adjustment_date or date.today()
        total_amount = abs(adjustment.quantity_change) * (product.cost_price or 0)

        # Use product's branch or find first available branch
        branch_id_to_use = adjustment.branch_id or product.branch_id
        if not branch_id_to_use:
            from app.models.branch import Branch
            first_branch = db.query(Branch).first()
            if first_branch:
                branch_id_to_use = first_branch.id

        accounting_entry = AccountingEntry(
            date_prepared=adjustment_date,
            date_posted=adjustment_date,
            particulars=f"Inventory Adjustment - {adjustment.adjustment_type}: {product.name}",
            book=f"INV-ADJ-{adjustment.adjustment_type.upper()}",
            status='posted',
            branch_id=branch_id_to_use
        )
        db.add(accounting_entry)
        db.flush()

        # Get inventory account
        inventory_account = product.accounting_code
        if not inventory_account:
            # Try to find a default inventory account
            inventory_account = db.query(AccountingCode).filter(
                AccountingCode.account_type == 'Asset',
                AccountingCode.name.ilike('%inventory%')
            ).first()

        if inventory_account and total_amount > 0:
            # Create journal entries based on adjustment type
            if adjustment.quantity_change > 0:
                # Increase inventory: Dr. Inventory, Cr. Inventory Gain/Opening Stock
                journal_entry1 = JournalEntry(
                    accounting_entry_id=accounting_entry.id,
                    accounting_code_id=inventory_account.id,
                    entry_type='debit',
                    debit_amount=total_amount,
                    description=f"Inventory increase - {adjustment.reason}",
                    date=adjustment_date,
                    date_posted=adjustment_date,
                    branch_id=accounting_entry.branch_id
                )
                db.add(journal_entry1)

                # Credit side - find or use a gain account
                gain_account = db.query(AccountingCode).filter(
                    AccountingCode.account_type == 'Income',
                    AccountingCode.name.ilike('%inventory gain%')
                ).first()

                if not gain_account:
                    # Use other income account
                    gain_account = db.query(AccountingCode).filter(
                        AccountingCode.account_type == 'Income',
                        AccountingCode.name.ilike('%other income%')
                    ).first()

                if gain_account:
                    journal_entry2 = JournalEntry(
                        accounting_entry_id=accounting_entry.id,
                        accounting_code_id=gain_account.id,
                        entry_type='credit',
                        credit_amount=total_amount,
                        description=f"Inventory gain - {adjustment.reason}",
                        date=adjustment_date,
                        date_posted=adjustment_date,
                        branch_id=accounting_entry.branch_id
                    )
                    db.add(journal_entry2)
            else:
                # Decrease inventory: Dr. Loss/Expense, Cr. Inventory
                loss_account = None

                if adjustment.adjustment_type == 'damage':
                    loss_account = db.query(AccountingCode).filter(
                        AccountingCode.account_type == 'Expense',
                        AccountingCode.name.ilike('%damage%')
                    ).first()
                elif adjustment.adjustment_type == 'theft':
                    loss_account = db.query(AccountingCode).filter(
                        AccountingCode.account_type == 'Expense',
                        AccountingCode.name.ilike('%theft%')
                    ).first()

                if not loss_account:
                    # Use inventory loss or other expense
                    loss_account = db.query(AccountingCode).filter(
                        AccountingCode.account_type == 'Expense',
                        AccountingCode.name.ilike('%inventory loss%')
                    ).first()

                if not loss_account:
                    loss_account = db.query(AccountingCode).filter(
                        AccountingCode.account_type == 'Expense',
                        AccountingCode.name.ilike('%other expense%')
                    ).first()

                if loss_account:
                    journal_entry1 = JournalEntry(
                        accounting_entry_id=accounting_entry.id,
                        accounting_code_id=loss_account.id,
                        entry_type='debit',
                        debit_amount=total_amount,
                        description=f"Inventory loss - {adjustment.adjustment_type} - {adjustment.reason}",
                        date=adjustment_date,
                        date_posted=adjustment_date,
                        branch_id=accounting_entry.branch_id
                    )
                    db.add(journal_entry1)

                journal_entry2 = JournalEntry(
                    accounting_entry_id=accounting_entry.id,
                    accounting_code_id=inventory_account.id,
                    entry_type='credit',
                    credit_amount=total_amount,
                    description=f"Inventory reduction - {adjustment.reason}",
                    date=adjustment_date,
                    date_posted=adjustment_date,
                    branch_id=accounting_entry.branch_id
                )
                db.add(journal_entry2)

        # Create inventory adjustment record
        db_adjustment = InventoryAdjustment(
            product_id=adjustment.product_id,
            adjustment_date=adjustment_date,
            quantity=abs(adjustment.quantity_change),
            reason=adjustment.reason,
            total_amount=total_amount,
            accounting_entry_type='inventory_adjustment',
            accounting_entry_id=accounting_entry.id,
            branch_id=branch_id_to_use,
            adjustment_type=adjustment.adjustment_type,
            previous_quantity=previous_quantity,
            new_quantity=new_quantity,
            unit_cost=product.cost_price,
            notes=adjustment.notes
        )
        db.add(db_adjustment)

        db.commit()
        db.refresh(db_adjustment)

        return db_adjustment

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating adjustment: {str(e)}")
