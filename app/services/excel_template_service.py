from __future__ import annotations

import mimetypes
import uuid
from io import BytesIO
from pathlib import Path
from typing import List, Optional

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session, joinedload

from app.core.config import settings
from app.models.excel_template import ExcelTemplate


class ExcelTemplateService:
    """Service layer for managing Excel templates."""

    ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
    ALLOWED_CONTENT_TYPES = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
    }

    def __init__(self, db: Session):
        self.db = db
        self.max_file_size = settings.max_file_size
        self.template_dir = Path("app/static/uploads/excel-templates")
        self.template_dir.mkdir(parents=True, exist_ok=True)

    async def upload_template(
        self,
        file: UploadFile,
        name: Optional[str] = None,
        description: Optional[str] = None,
        category: Optional[str] = None,
        uploaded_by: Optional[str] = None,
    ) -> ExcelTemplate:
        """Persist a new Excel template to storage and the database."""

        if not file.filename:
            raise HTTPException(status_code=400, detail="Uploaded file is missing a filename")

        extension = Path(file.filename).suffix.lower()
        if extension not in self.ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail="Invalid template format. Only .xlsx and .xlsm files are supported.",
            )

        contents = await file.read()
        if not contents:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")

        if len(contents) > self.max_file_size:
            raise HTTPException(
                status_code=400,
                detail=f"File too large. Maximum allowed size is {self.max_file_size // (1024 * 1024)} MB.",
            )

        content_type = (file.content_type or "").split(";")[0]
        if content_type and content_type not in self.ALLOWED_CONTENT_TYPES:
            raise HTTPException(
                status_code=400,
                detail="Invalid content type. Please upload a valid Excel workbook.",
            )

        # Validate workbook and extract metadata before saving
        metadata = self._extract_workbook_metadata(contents)

        stored_filename = f"excel_template_{uuid.uuid4().hex}{extension}"
        file_path = self.template_dir / stored_filename
        with open(file_path, "wb") as output_file:
            output_file.write(contents)

        template = ExcelTemplate(
            name=name or Path(file.filename).stem,
            description=description,
            category=category,
            original_filename=file.filename,
            stored_filename=stored_filename,
            file_path=str(file_path),
            public_url=f"/static/uploads/excel-templates/{stored_filename}",
            file_size=len(contents),
            content_type=content_type or self._guess_content_type(extension),
            workbook_metadata=metadata,
            uploaded_by=uploaded_by,
        )

        document_properties = (metadata or {}).get("document_properties") or {}
        if "version" in document_properties and document_properties["version"]:
            template.version = str(document_properties["version"])

        self.db.add(template)
        self.db.commit()
        self.db.refresh(template)
        return self.get_template(template.id)

    def list_templates(self) -> List[ExcelTemplate]:
        """Return all stored templates ordered by creation time (newest first)."""
        return (
            self.db.query(ExcelTemplate)
            .options(
                joinedload(ExcelTemplate.uploaded_by_user).joinedload("branch")
            )
            .order_by(ExcelTemplate.created_at.desc())
            .all()
        )

    def get_template(self, template_id: str) -> ExcelTemplate:
        """Retrieve a template or raise 404 if it does not exist."""
        template = (
            self.db.query(ExcelTemplate)
            .options(
                joinedload(ExcelTemplate.uploaded_by_user).joinedload("branch")
            )
            .filter(ExcelTemplate.id == template_id)
            .first()
        )
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        return template

    def delete_template(self, template_id: str) -> None:
        """Remove a template and its stored file."""
        template = self.get_template(template_id)
        file_path = Path(template.file_path)
        if file_path.exists():
            try:
                file_path.unlink()
            except OSError:
                # Ignore file system errors; the database record will still be removed
                pass

        self.db.delete(template)
        self.db.commit()

    def build_download_url(self, template_id: str) -> str:
        template = self.get_template(template_id)
        return f"/api/v1/excel-templates/{template.id}/download"

    def _extract_workbook_metadata(self, content: bytes):
        """Parse the workbook to collect structural metadata for field mapping."""
        metadata = {
            "sheet_names": [],
            "named_ranges": [],
            "sample_headers": {},
            "document_properties": {},
        }

        workbook = None
        try:
            workbook = load_workbook(
                filename=BytesIO(content),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
            metadata["sheet_names"] = list(workbook.sheetnames)

            # Named ranges / defined names
            defined_names = workbook.defined_names.definedName if workbook.defined_names else []
            metadata["named_ranges"] = sorted(
                {
                    dn.name
                    for dn in defined_names
                    if getattr(dn, "name", None)
                }
            )

            for sheet_name in metadata["sheet_names"]:
                worksheet = workbook[sheet_name]
                sample_row: List[Optional[str]] = []
                for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
                    sample_row = [self._stringify_cell(value) for value in row]
                    break
                metadata["sample_headers"][sheet_name] = sample_row

            document_properties = {}
            props = getattr(workbook, "properties", None)
            if props:
                for attr in (
                    "title",
                    "subject",
                    "creator",
                    "lastModifiedBy",
                    "created",
                    "modified",
                    "keywords",
                    "category",
                ):
                    value = getattr(props, attr, None)
                    if value is None:
                        continue
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()
                    document_properties[attr] = value
            metadata["document_properties"] = document_properties

            return metadata
        except InvalidFileException as exc:
            raise HTTPException(status_code=400, detail="Uploaded file is not a valid Excel workbook") from exc
        except Exception as exc:  # pylint: disable=broad-except
            raise HTTPException(status_code=400, detail=f"Unable to read workbook: {exc}") from exc
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:  # pragma: no cover - best effort cleanup
                    pass

    @staticmethod
    def _stringify_cell(value):
        if value is None:
            return None
        if isinstance(value, (str, int, float)):
            return value
        return str(value)

    @staticmethod
    def _guess_content_type(extension: str) -> str:
        content_type, _ = mimetypes.guess_type(f"placeholder{extension}")
        return content_type or "application/octet-stream"
